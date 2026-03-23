"""
Bispecific export — CSV, XLSX, zone-specific PDBs, and dual PML scripts.
"""

import csv
import logging
from pathlib import Path
from typing import Dict, List

from epitope_pipeline.config import PALETTE, DUAL_PML_GAP_A
from epitope_pipeline.export import (
    export_annotated_pdb,
    export_blast_details,
    generate_membrane_cgo_pml,
    generate_shared_membrane_cgo_pml,
)

logger = logging.getLogger("epitope_pipeline")


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def export_bispecific_all(run_dir, pair_results, zone_results,
                          structures, membranes):
    """
    Generate all bispecific-specific output files.

    Args:
        run_dir: Path to the run directory.
        pair_results: List of BispecificPairResult.
        zone_results: Dict {(uid, zone): TargetZoneResult}.
        structures: Dict {uid: StructureResult}.
        membranes: Dict {uid: MembraneAnnotation}.
    """
    run_dir = Path(run_dir)
    pymol_dir = run_dir / "Structures"
    pymol_dir.mkdir(exist_ok=True)
    zone_sessions_dir = pymol_dir / "zone_sessions"
    zone_sessions_dir.mkdir(exist_ok=True)

    # 1. CSV
    export_bispecific_csv(run_dir, pair_results)

    # 2. XLSX
    export_bispecific_xlsx(run_dir, pair_results)

    # 3. Zone-specific annotated PDBs + PMLs → pymol/zone_sessions/
    exported_zones = set()
    for key, zr in zone_results.items():
        uid, zone = key
        if (uid, zone) in exported_zones:
            continue
        export_zone_annotated_pdb(run_dir, zr)
        exported_zones.add((uid, zone))

    # 4. BLAST detail files → blast/
    exported_blast = set()
    for key, zr in zone_results.items():
        uid, zone = key
        if uid in exported_blast:
            continue
        if zr.specificity_result:
            export_blast_details(run_dir / "Supplementary Files" / "BLAST", zr.target, zr.specificity_result)
            exported_blast.add(uid)

    # 5. Dual PML scripts (one per orientation) → pymol/
    for pr in pair_results:
        for orientation in (pr.orientation_ab, pr.orientation_ba):
            export_bispecific_pml(
                run_dir, pr, orientation, structures, membranes,
            )

    logger.info("  Bispecific outputs written to: %s", run_dir)


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

def export_bispecific_csv(run_dir, pair_results):
    """Write bispecific_pairs.csv (hidden)."""
    csv_path = Path(run_dir) / ".bispecific_pairs.csv"
    rows = []

    for pr in pair_results:
        for orientation in (pr.orientation_ab, pr.orientation_ba):
            distal_gene = orientation.distal_zone.target.gene_name if orientation.distal_zone else "—"
            proximal_gene = orientation.proximal_zone.target.gene_name if orientation.proximal_zone else "—"

            distal_best = 0.0
            distal_area = 0.0
            distal_patch_id = "—"
            if orientation.distal_zone and orientation.distal_zone.scores:
                s = orientation.distal_zone.scores[0]
                distal_best = s.composite_score
                distal_area = s.patch_area_a2
                distal_patch_id = s.patch_id

            proximal_best = 0.0
            proximal_area = 0.0
            proximal_patch_id = "—"
            if orientation.proximal_zone and orientation.proximal_zone.scores:
                s = orientation.proximal_zone.scores[0]
                proximal_best = s.composite_score
                proximal_area = s.patch_area_a2
                proximal_patch_id = s.patch_id

            rows.append({
                "pair": "{} x {}".format(pr.target_a.gene_name, pr.target_b.gene_name),
                "target_a": pr.target_a.gene_name,
                "target_b": pr.target_b.gene_name,
                "orientation": orientation.label,
                "distal_target": distal_gene,
                "proximal_target": proximal_gene,
                "distal_best_patch": distal_patch_id,
                "distal_best_score": round(distal_best, 4),
                "distal_best_area_A2": round(distal_area, 1),
                "proximal_best_patch": proximal_patch_id,
                "proximal_best_score": round(proximal_best, 4),
                "proximal_best_area_A2": round(proximal_area, 1),
                "orientation_score": round(orientation.orientation_score, 4),
                "is_valid": orientation.is_valid,
                "dual_valid": pr.both_valid,
                "flexibility_bonus": pr.both_valid,
                "final_pair_score": round(pr.final_pair_score, 4),
            })

    if rows:
        fieldnames = list(rows[0].keys())
        with open(csv_path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
        logger.info("  Wrote %s (%d rows)", csv_path.name, len(rows))


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------

def export_bispecific_xlsx(run_dir, pair_results):
    """Write bispecific_pairs.xlsx with Cartography styling."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    supp_dir = Path(run_dir) / "Supplementary Files"
    supp_dir.mkdir(exist_ok=True)
    xlsx_path = supp_dir / "bispecific_pairs.xlsx"
    wb = Workbook()

    # Styles
    header_fill = PatternFill(start_color="28154C", end_color="28154C", fill_type="solid")
    header_font = Font(name="Consolas", bold=True, color="FFFFFF", size=10)
    data_font = Font(name="Consolas", size=10)
    score_font = Font(name="Consolas", size=10, bold=True)
    valid_fill = PatternFill(start_color="6BC291", end_color="6BC291", fill_type="solid")
    invalid_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    dual_fill = PatternFill(start_color="18B5CB", end_color="18B5CB", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # === Sheet 1: Pair Summary ===
    ws = wb.active
    ws.title = "Pair Summary"

    headers = [
        "Pair", "Target A", "Target B", "Final Pair Score",
        "Dual-Valid", "Best Orientation",
        "Distal Target", "Distal Best Score",
        "Proximal Target", "Proximal Best Score",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, pr in enumerate(pair_results, 2):
        best = pr.best_orientation
        distal_gene = best.distal_zone.target.gene_name if best.distal_zone else "—"
        proximal_gene = best.proximal_zone.target.gene_name if best.proximal_zone else "—"
        distal_score = best.distal_zone.best_score if best.distal_zone else 0.0
        proximal_score = best.proximal_zone.best_score if best.proximal_zone else 0.0

        values = [
            "{} x {}".format(pr.target_a.gene_name, pr.target_b.gene_name),
            pr.target_a.gene_name,
            pr.target_b.gene_name,
            round(pr.final_pair_score, 4),
            "Yes" if pr.both_valid else "No",
            best.label,
            distal_gene,
            round(distal_score, 4),
            proximal_gene,
            round(proximal_score, 4),
        ]

        if pr.both_valid:
            fill = dual_fill
        elif pr.final_pair_score > 0:
            fill = valid_fill
        else:
            fill = invalid_fill

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = score_font if col == 4 else data_font
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # === Sheet 2: All Orientations ===
    ws2 = wb.create_sheet("Orientation Detail")

    orient_headers = [
        "Pair", "Orientation", "Valid",
        "Distal Target", "Distal Score", "Distal Patches", "Distal Area (A2)",
        "Proximal Target", "Proximal Score", "Proximal Patches", "Proximal Area (A2)",
        "Orientation Score",
    ]
    for col, h in enumerate(orient_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row_idx = 2
    for pr in pair_results:
        for orientation in (pr.orientation_ab, pr.orientation_ba):
            dz = orientation.distal_zone
            pz = orientation.proximal_zone

            values = [
                "{} x {}".format(pr.target_a.gene_name, pr.target_b.gene_name),
                orientation.label,
                "Yes" if orientation.is_valid else "No",
                dz.target.gene_name if dz else "—",
                round(dz.best_score, 4) if dz else 0.0,
                len(dz.scores) if dz else 0,
                round(sum(s.patch_area_a2 for s in dz.scores), 1) if dz and dz.scores else 0.0,
                pz.target.gene_name if pz else "—",
                round(pz.best_score, 4) if pz else 0.0,
                len(pz.scores) if pz else 0,
                round(sum(s.patch_area_a2 for s in pz.scores), 1) if pz and pz.scores else 0.0,
                round(orientation.orientation_score, 4),
            ]

            fill = valid_fill if orientation.is_valid else invalid_fill
            for col, val in enumerate(values, 1):
                cell = ws2.cell(row=row_idx, column=col, value=val)
                cell.font = data_font
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
            row_idx += 1

    for col in range(1, len(orient_headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    wb.save(str(xlsx_path))
    logger.info("  Wrote %s", xlsx_path.name)


# ---------------------------------------------------------------------------
# Zone-specific annotated PDBs
# ---------------------------------------------------------------------------

def export_zone_annotated_pdb(run_dir, zone_result):
    """
    Write {gene}_{zone}_epitope.pdb with zone-specific B-factor tiers.

    Reuses the existing export_annotated_pdb() from export.py but writes
    to a zone-specific filename.
    """
    target = zone_result.target
    gene = target.gene_name.lower()
    zone = zone_result.zone

    # Temporarily rename gene_name to get zone-specific filenames
    # from export_annotated_pdb (which uses target.gene_name.lower())
    original_name = target.gene_name
    target.gene_name = "{}_{}".format(original_name, zone)

    try:
        export_annotated_pdb(
            run_dir=str(run_dir),
            target=target,
            structure=zone_result.structure,
            membrane=zone_result.membrane,
            spatial_filter=zone_result.spatial_filter,
            surface_analysis=zone_result.surface_analysis,
            conservation_result=zone_result.conservation_result,
            scores=zone_result.scores,
            pdb_subdir="Structures/zone_sessions",
        )
    finally:
        # Restore original name
        target.gene_name = original_name


# ---------------------------------------------------------------------------
# Dual PML script
# ---------------------------------------------------------------------------

def export_bispecific_pml(run_dir, pair_result, orientation,
                          structures, membranes):
    """
    Write a dual-target PyMOL script with both structures side-by-side.
    """
    run_dir = Path(run_dir)
    pml_dir = run_dir / "Structures"
    pml_dir.mkdir(parents=True, exist_ok=True)

    dz = orientation.distal_zone
    pz = orientation.proximal_zone
    distal_gene = dz.target.gene_name.lower()
    proximal_gene = pz.target.gene_name.lower()

    # PyMOL object names — short gene names for clean panel display
    distal_pdb = dz.target.gene_name
    proximal_pdb = pz.target.gene_name
    # File names on disk (dot-prefixed, hidden)
    distal_pdb_file = "{}_distal_epitope".format(distal_gene)
    proximal_pdb_file = "{}_proximal_epitope".format(proximal_gene)

    pml_path = pml_dir / "{}.pml".format(orientation.label.lower())

    # Build residue sets for each target's patches
    distal_patches = {}
    for s in dz.scores:
        distal_patches[s.patch_id] = sorted(s.patch.residue_numbers)

    proximal_patches = {}
    for s in pz.scores:
        proximal_patches[s.patch_id] = sorted(s.patch.residue_numbers)

    # Helper: format residue numbers for PyMOL resi selection
    def resi_str(resnums):
        if not resnums:
            return "none"
        nums = sorted(resnums)
        ranges = []
        start = nums[0]
        end = nums[0]
        for n in nums[1:]:
            if n == end + 1:
                end = n
            else:
                ranges.append("{}-{}".format(start, end) if end > start else str(start))
                start = end = n
        ranges.append("{}-{}".format(start, end) if end > start else str(start))
        return "+".join(ranges)

    # Determine chains
    distal_chain = dz.structure.chain_id
    proximal_chain = pz.structure.chain_id

    gap = DUAL_PML_GAP_A

    # --- Compute translation and bilayer geometry ---
    # If both zone PDBs have rotated coordinates (single-pass TM with
    # ecto-axis → Y-up, anchor at origin), use simple X translation and
    # a horizontal bilayer. Otherwise fall back to membrane-normal alignment.
    import numpy as np

    def _is_rotated(zone):
        sf = zone.spatial_filter
        return (sf is not None
                and getattr(sf, 'anchor_resnum', None) is not None
                and getattr(sf, 'farthest_resnum', None) is not None)

    both_rotated = _is_rotated(dz) and _is_rotated(pz)

    if both_rotated:
        # Both PDBs rotated: anchors at origin, ecto-axis = Y-up.
        # Simple X translation, horizontal bilayer.
        translate_vec = np.array([gap, 0.0, 0.0])
        avg_ht = (dz.membrane.membrane_half_thickness +
                  pz.membrane.membrane_half_thickness) / 2.0
        cgo_center = np.array([gap / 2.0, -avg_ht, 0.0])
        cgo_normal = np.array([0.0, 1.0, 0.0])
    else:
        # Membrane-normal-based alignment (multi-pass / GPI / mixed)
        normal_d = np.array(dz.membrane.membrane_normal)
        normal_p = np.array(pz.membrane.membrane_normal)
        avg_normal = normal_d + normal_p
        n_len = np.linalg.norm(avg_normal)
        up = avg_normal / n_len if n_len > 0 else np.array([0., 1., 0.])

        right = np.array([1.0, 0.0, 0.0])
        right = right - np.dot(right, up) * up
        if np.linalg.norm(right) < 0.01:
            right = np.array([0.0, 0.0, 1.0])
            right = right - np.dot(right, up) * up
        right = right / np.linalg.norm(right)
        fwd = np.cross(right, up)
        fwd = fwd / np.linalg.norm(fwd)

        ref_d = np.array(dz.membrane.membrane_center)
        ref_p = np.array(pz.membrane.membrane_center)
        vertical_offset = (np.dot(ref_d, up) - np.dot(ref_p, up)) * up
        translate_vec = right * gap + vertical_offset
        cgo_center = None
        cgo_normal = None

    lines = [
        "# ================================================================",
        "# BISPECIFIC VISUALIZATION",
        "# {} (distal) x {} (proximal)".format(
            dz.target.gene_name, pz.target.gene_name),
        "# Pair score: {:.3f}{}".format(
            pair_result.final_pair_score,
            " [DUAL-VALID, {:.0f}% bonus]".format(
                (pair_result.final_pair_score / orientation.orientation_score - 1) * 100
            ) if pair_result.both_valid and orientation.orientation_score > 0 else "",
        ),
        "# ================================================================",
        "",
        "# --- Cartography palette ---",
        "set_color carto_gray,       [0.827, 0.827, 0.827]",
        "set_color carto_palegreen,  [0.878, 0.953, 0.859]",
        "set_color carto_green,      [0.420, 0.761, 0.569]",
        "set_color carto_teal,       [0.094, 0.710, 0.796]",
        "set_color carto_blue,       [0.180, 0.584, 0.824]",
        "set_color carto_purple,     [0.157, 0.082, 0.298]",
        "",
        "# --- Load both structures ---",
        "load zone_sessions/.{}.pdb, {}".format(distal_pdb_file, distal_pdb),
        "load zone_sessions/.{}.pdb, {}".format(proximal_pdb_file, proximal_pdb),
        "",
        "# --- Translate proximal target for side-by-side view ---",
        "translate [{:.2f}, {:.2f}, {:.2f}], {}".format(
            translate_vec[0], translate_vec[1], translate_vec[2], proximal_pdb),
        "",
        "# --- Base display ---",
        "hide everything",
        "show cartoon, {}".format(distal_pdb),
        "show cartoon, {}".format(proximal_pdb),
        "color white, {}".format(distal_pdb),
        "color white, {}".format(proximal_pdb),
        "set cartoon_transparency, 0.3, {}".format(distal_pdb),
        "set cartoon_transparency, 0.3, {}".format(proximal_pdb),
        "",
    ]

    # --- Per-target topology handling, filter selections, and patches ---
    # Helper: generate all selections for one zone (mirrors _write_pymol_script logic)
    def _zone_lines(zone, pdb_obj, ch, role, patches_dict):
        """Generate PML lines for topology + filter tiers + patches for one zone."""
        zlines = []
        gene = zone.target.gene_name
        # Short suffix for selection uniqueness (_D or _P) — group name provides full context
        sx = "_D" if role == "Distal" else "_P"
        prefix = "{}_{}".format(gene, role)  # for group names

        # --- Topology handling (multi-pass vs single-pass/GPI) ---
        is_multipass = (zone.membrane and
                        zone.membrane.topology_type == "multi_pass")
        if is_multipass:
            zlines.append("# --- Multi-pass TM: {} — full structure with topology coloring ---".format(gene))
            zlines.append("set cartoon_transparency, 0.0, {}".format(pdb_obj))
            zlines.append("")

            tm_residues = set()
            for seg_start, seg_end in zone.membrane.tm_segments:
                for r in range(seg_start, seg_end + 1):
                    tm_residues.add(r)
            if tm_residues:
                zlines.append("select TM_Helices{}, {} and chain {} and resi {}".format(
                    sx, pdb_obj, ch, resi_str(tm_residues)))
                zlines.append("color carto_purple, TM_Helices{}".format(sx))
                zlines.append("")

            ec_residues = set()
            ic_residues = set()
            for resnum, topo in zone.membrane.residue_topology.items():
                if topo == "extracellular":
                    ec_residues.add(resnum)
                elif topo == "intracellular":
                    ic_residues.add(resnum)
            if ec_residues:
                zlines.append("select ECD_Loops{}, {} and chain {} and resi {}".format(
                    sx, pdb_obj, ch, resi_str(ec_residues)))
                zlines.append("color carto_teal, ECD_Loops{}".format(sx))
                zlines.append("")
            if ic_residues:
                zlines.append("select ICD_Loops{}, {} and chain {} and resi {}".format(
                    sx, pdb_obj, ch, resi_str(ic_residues)))
                zlines.append("color carto_palegreen, ICD_Loops{}".format(sx))
                zlines.append("disable ICD_Loops{}".format(sx))
                zlines.append("")
        else:
            # Single-pass / GPI: hide cytoplasmic
            cyto = set()
            if zone.membrane:
                for resnum, topo in zone.membrane.residue_topology.items():
                    if topo == "intracellular":
                        cyto.add(resnum)
            if cyto:
                zlines.append("# --- Hide cytoplasmic: {} ---".format(gene))
                zlines.append("select _cyto{}, {} and chain {} and resi {}".format(
                    sx, pdb_obj, ch, resi_str(cyto)))
                zlines.append("hide cartoon, _cyto{}".format(sx))
                zlines.append("disable _cyto{}".format(sx))
                zlines.append("")

        # --- Build filter tier sets (mirrors _write_pymol_script) ---
        patch_set = set()
        for s in zone.scores:
            for resnum in s.patch.residue_numbers:
                patch_set.add(resnum)

        dist_qualified = set()
        if zone.spatial_filter:
            dist_qualified = set(zone.spatial_filter.qualifying_residues)

        surf_exposed = set()
        if zone.surface_analysis:
            surf_exposed = set(zone.surface_analysis.exposed_residues)

        cyno_conserved = set()
        if zone.conservation_result:
            for resnum, is_conserved in zone.conservation_result.residue_conservation.items():
                if is_conserved is True:
                    cyno_conserved.add(resnum)

        dist_thresh = zone.spatial_filter.min_distance_threshold if zone.spatial_filter else 0

        conserved_surf = (cyno_conserved & surf_exposed) - patch_set
        surf_only = surf_exposed - cyno_conserved - patch_set
        dist_only = dist_qualified - surf_exposed - patch_set

        all_ec = set()
        if zone.membrane:
            for resnum, topo in zone.membrane.residue_topology.items():
                if topo == "extracellular":
                    all_ec.add(resnum)
        ec_near = all_ec - dist_qualified - patch_set

        # --- Epitope patches → grouped under "{prefix}_Epitopes" ---
        zone_label = ">= {}A".format(int(dist_thresh)) if role == "Distal" else "<= {}A".format(
            int(zone.spatial_filter.max_distance_threshold) if zone.spatial_filter and zone.spatial_filter.max_distance_threshold else "?")
        zlines.append("# =============================================")
        zlines.append("# {} TARGET: {} ({})".format(role.upper(), gene, zone_label))
        zlines.append("# =============================================")
        zlines.append("")

        patch_names = []
        for pid, resnums in patches_dict.items():
            sel_name = "Patch_{}{}".format(pid, sx)
            patch_names.append(sel_name)
            zlines.append("select {}, {} and chain {} and resi {}".format(
                sel_name, pdb_obj, ch, resi_str(resnums)))
            zlines.append("color carto_green, {}".format(sel_name))
            zlines.append("show surface, {}".format(sel_name))
            zlines.append("set transparency, 0.3, {}".format(sel_name))
            zlines.append("set cartoon_transparency, 0.0, {}".format(sel_name))
            zlines.append("")

        if patch_set:
            combined_name = "Combined{}".format(sx)
            zlines.append("select {}, {} and chain {} and resi {}".format(
                combined_name, pdb_obj, ch, resi_str(patch_set)))
            patch_names.append(combined_name)
            zlines.append("")

        epi_group = "Epitopes{}".format(sx)
        if patch_names:
            zlines.append("group {}, {}".format(epi_group, " ".join(patch_names)))
            zlines.append("group {}, close".format(epi_group))
            zlines.append("")

        # --- Filter tiers → grouped under "Filters" (closed, but clickable) ---
        filter_names = []

        if ec_near:
            sel = "EC_Near_Membrane{}".format(sx)
            zlines.append("select {}, {} and chain {} and resi {}".format(
                sel, pdb_obj, ch, resi_str(ec_near)))
            filter_names.append(sel)
            zlines.append("")

        if dist_only:
            sel = "Distant_Buried{}".format(sx)
            zlines.append("select {}, {} and chain {} and resi {}".format(
                sel, pdb_obj, ch, resi_str(dist_only)))
            filter_names.append(sel)
            zlines.append("")

        if surf_only:
            sel = "Exposed_Not_Conserved{}".format(sx)
            zlines.append("select {}, {} and chain {} and resi {}".format(
                sel, pdb_obj, ch, resi_str(surf_only)))
            filter_names.append(sel)
            zlines.append("")

        if conserved_surf:
            sel = "Conserved_Not_Patched{}".format(sx)
            zlines.append("select {}, {} and chain {} and resi {}".format(
                sel, pdb_obj, ch, resi_str(conserved_surf)))
            filter_names.append(sel)
            zlines.append("")

        all_dist_pass = dist_qualified | patch_set
        all_surf_pass = (surf_exposed & dist_qualified) | patch_set
        all_cons_pass = (cyno_conserved & surf_exposed & dist_qualified) | patch_set

        if all_dist_pass:
            sel = "Cumul_Distance_Pass{}".format(sx)
            zlines.append("select {}, {} and chain {} and resi {}".format(
                sel, pdb_obj, ch, resi_str(all_dist_pass)))
            filter_names.append(sel)
            zlines.append("")
        if all_surf_pass:
            sel = "Cumul_Surface_Pass{}".format(sx)
            zlines.append("select {}, {} and chain {} and resi {}".format(
                sel, pdb_obj, ch, resi_str(all_surf_pass)))
            filter_names.append(sel)
            zlines.append("")
        if all_cons_pass:
            sel = "Cumul_Conservation_Pass{}".format(sx)
            zlines.append("select {}, {} and chain {} and resi {}".format(
                sel, pdb_obj, ch, resi_str(all_cons_pass)))
            filter_names.append(sel)
            zlines.append("")

        filt_group = "Filters{}".format(sx)
        if filter_names:
            zlines.append("group {}, {}".format(filt_group, " ".join(filter_names)))
            zlines.append("group {}, close".format(filt_group))
            zlines.append("")

        # Top-level group: gene_Role containing PDB object + epitopes + filters
        top_group = "{}_{}".format(gene, role)
        sub_items = [pdb_obj]  # include the structure object
        if patch_names:
            sub_items.append(epi_group)
        if filter_names:
            sub_items.append(filt_group)
        zlines.append("group {}, {}".format(top_group, " ".join(sub_items)))
        zlines.append("group {}, close".format(top_group))
        zlines.append("")

        return zlines

    # Generate selections for distal target
    lines.extend(_zone_lines(dz, distal_pdb, distal_chain, "Distal", distal_patches))

    # Generate selections for proximal target
    lines.extend(_zone_lines(pz, proximal_pdb, proximal_chain, "Proximal", proximal_patches))

    # Shared membrane bilayer CGO spanning both structures
    pdb_dir = run_dir / "Structures" / "zone_sessions"
    if dz.membrane and pz.membrane:
        distal_pdb_path = str(pdb_dir / ".{}.pdb".format(distal_pdb_file))
        proximal_pdb_path = str(pdb_dir / ".{}.pdb".format(proximal_pdb_file))
        lines.extend(generate_shared_membrane_cgo_pml(
            membrane_a=dz.membrane, pdb_path_a=distal_pdb_path, chain_a=distal_chain,
            membrane_b=pz.membrane, pdb_path_b=proximal_pdb_path, chain_b=proximal_chain,
            translate_b=translate_vec,
            normal_override=cgo_normal,
            center_override=cgo_center,
        ))
    elif dz.membrane:
        d_pdb_path = str(pdb_dir / ".{}.pdb".format(distal_pdb_file))
        lines.extend(generate_membrane_cgo_pml(
            dz.membrane, d_pdb_path, distal_chain,
            normal_override=cgo_normal,
            center_override=cgo_center,
        ))
    elif pz.membrane:
        p_pdb_path = str(pdb_dir / ".{}.pdb".format(proximal_pdb_file))
        lines.extend(generate_membrane_cgo_pml(
            pz.membrane, p_pdb_path, proximal_chain,
            translate_vec=translate_vec,
            normal_override=cgo_normal,
            center_override=cgo_center,
        ))

    # Labels
    lines.extend([
        "# --- Labels ---",
        "set label_size, 16",
        "set label_color, black",
        "",
    ])

    # --- View ---
    if both_rotated:
        # Both PDBs rotated: ecto-axis = Y-up. Identity view + zoom.
        lines.extend([
            "# --- Ectodomain-axis-aligned view (coordinates rotated) ---",
            "deselect",
            "set_view (\\",
            "    1.000000, 0.000000, 0.000000,\\",
            "    0.000000, 1.000000, 0.000000,\\",
            "    0.000000, 0.000000, 1.000000,\\",
            "    0.000000, 0.000000, -400.000000,\\",
            "    {:.2f}, 0.00, 0.00,\\".format(gap / 2.0),
            "    100.000000, 900.000000, -20.000000)",
            "",
            "zoom",
            "",
        ])
    else:
        # Membrane-normal-based view with ectodomain-up check
        center_d_mem = np.array(dz.membrane.membrane_center)
        center_p_mem = np.array(pz.membrane.membrane_center) + translate_vec
        view_center = (center_d_mem + center_p_mem) / 2.0
        lines.extend([
            "# --- Membrane-aligned view ---",
            "deselect",
            "set_view (\\",
            "    {:.6f}, {:.6f}, {:.6f},\\".format(right[0], right[1], right[2]),
            "    {:.6f}, {:.6f}, {:.6f},\\".format(up[0], up[1], up[2]),
            "    {:.6f}, {:.6f}, {:.6f},\\".format(fwd[0], fwd[1], fwd[2]),
            "    0.000000, 0.000000, -400.000000,\\",
            "    {:.2f}, {:.2f}, {:.2f},\\".format(
                view_center[0], view_center[1], view_center[2]),
            "    100.000000, 900.000000, -20.000000)",
            "",
            "# Ensure ectodomain faces UP",
            "python",
            "from pymol import cmd",
            "v = cmd.get_view()",
            "com = cmd.centerofmass('visible')",
            "mcx, mcy, mcz = {:.2f}, {:.2f}, {:.2f}".format(
                view_center[0], view_center[1], view_center[2]),
            "dx = com[0] - mcx",
            "dy = com[1] - mcy",
            "dz2 = com[2] - mcz",
            "sy_com = v[3]*dx + v[4]*dy + v[5]*dz2",
            "if sy_com < 0:",
            "    cmd.turn('x', 180)",
            "python end",
            "",
            "zoom",
            "",
        ])

    # Background settings (match single-target)
    lines.extend([
        "set ray_opaque_background, off",
        "bg_color black",
        "",
    ])

    with open(pml_path, "w") as f:
        f.write("\n".join(lines) + "\n")

    logger.info("  Wrote dual PML: %s", pml_path.name)
