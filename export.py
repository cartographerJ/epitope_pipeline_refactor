"""
Step 9: Output Generation — produce all output files for a pipeline run.

Generates:
  - epitope_candidates.csv — main results table
  - epitope_candidates.xlsx — color-formatted Excel workbook (Cartography palette)
  - annotated PDB files — B-factor column = epitope score (PyMOL/ChimeraX)
  - annotated FASTA files — headers with epitope patch residue ranges
  - per-residue annotation CSVs — detailed residue-level data
  - JSON annotations — full intermediate results for programmatic reuse
  - input_manifest.json — reproducibility record
"""

import csv
import json
import logging
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional

import numpy as np
from Bio.PDB import PDBParser, PDBIO, Select

from epitope_pipeline.config import (
    COLOR_CONSERVED,
    COLOR_EPITOPE_PATCH,
    COLOR_EXTRACELLULAR,
    COLOR_INTRACELLULAR,
    COLOR_MISMATCH,
    COLOR_TRANSMEMBRANE,
    PALETTE,
    SPECIFICITY_IDENTITY_THRESHOLD,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _rotation_matrix_align(from_vec, to_vec):
    """Rotation matrix that maps from_vec to to_vec (Rodrigues' formula)."""
    a = np.asarray(from_vec, dtype=float)
    b = np.asarray(to_vec, dtype=float)
    a = a / np.linalg.norm(a)
    b = b / np.linalg.norm(b)
    v = np.cross(a, b)
    c = np.dot(a, b)
    s = np.linalg.norm(v)
    if s < 1e-8:
        if c > 0:
            return np.eye(3)
        # 180-degree rotation: pick an arbitrary perpendicular axis
        perp = np.array([1.0, 0.0, 0.0]) if abs(a[0]) < 0.9 else np.array([0.0, 1.0, 0.0])
        perp = perp - np.dot(perp, a) * a
        perp = perp / np.linalg.norm(perp)
        return 2.0 * np.outer(perp, perp) - np.eye(3)
    K = np.array([
        [0, -v[2], v[1]],
        [v[2], 0, -v[0]],
        [-v[1], v[0], 0],
    ])
    return np.eye(3) + K + K @ K * (1.0 - c) / (s * s)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def export_all(
    run_dir,
    targets,
    epitope_scores,
    structures,
    membranes,
    spatial_filters,
    surface_analyses,
    conservation_results,
    specificity_results,
    target_metrics,
    parameters,
):
    """
    Generate all output files for a pipeline run.

    Args:
        run_dir: Path to the run output directory.
        targets: List of TargetInfo objects.
        epitope_scores: Dict {uniprot_id: [EpitopeScore, ...]}.
        structures: Dict {uniprot_id: StructureResult}.
        membranes: Dict {uniprot_id: MembraneAnnotation}.
        spatial_filters: Dict {uniprot_id: SpatialFilter}.
        surface_analyses: Dict {uniprot_id: SurfaceAnalysis}.
        conservation_results: Dict {uniprot_id: ConservationResult}.
        specificity_results: Dict {uniprot_id: SpecificityResult}.
        target_metrics: Dict {uniprot_id: metric_dict}.
        parameters: Dict of pipeline parameters used.
    """
    run_dir = Path(run_dir)
    run_dir.mkdir(parents=True, exist_ok=True)
    (run_dir / "Structures").mkdir(exist_ok=True)
    supp_dir = run_dir / "Supplementary Files"
    supp_dir.mkdir(exist_ok=True)
    (supp_dir / "Annotated Sequences").mkdir(exist_ok=True)
    (supp_dir / "Annotations").mkdir(exist_ok=True)

    # 1. Main results CSV
    export_results_csv(run_dir, targets, epitope_scores, target_metrics)

    # 2. Excel workbook → Supplementary Files/
    export_results_xlsx(supp_dir, targets, epitope_scores, target_metrics,
                        membranes, conservation_results)

    # 3. Annotated PDBs
    for target in targets:
        uid = target.uniprot_id
        if uid in structures and uid in membranes:
            export_annotated_pdb(
                run_dir, target, structures[uid], membranes[uid],
                spatial_filters.get(uid),
                surface_analyses.get(uid),
                conservation_results.get(uid),
                epitope_scores.get(uid, []),
            )

    # 4. Annotated sequences → Supplementary Files/
    for target in targets:
        uid = target.uniprot_id
        export_annotated_sequences(
            supp_dir, target,
            membranes.get(uid),
            spatial_filters.get(uid),
            surface_analyses.get(uid),
            conservation_results.get(uid),
            specificity_results.get(uid),
            epitope_scores.get(uid, []),
        )

    # 5. BLAST detail files → Supplementary Files/
    for target in targets:
        uid = target.uniprot_id
        if specificity_results.get(uid):
            export_blast_details(supp_dir / "BLAST", target, specificity_results[uid])

    # 6. JSON annotations → Supplementary Files/
    for target in targets:
        uid = target.uniprot_id
        export_annotation_json(
            supp_dir, target,
            structures.get(uid),
            membranes.get(uid),
            spatial_filters.get(uid),
            surface_analyses.get(uid),
            conservation_results.get(uid),
            specificity_results.get(uid),
            epitope_scores.get(uid, []),
            target_metrics.get(uid, {}),
        )

    # 7. Input manifest
    export_manifest(run_dir, targets, parameters)

    logger.info("All outputs written to: %s", run_dir)


# ---------------------------------------------------------------------------
# BLAST detail export
# ---------------------------------------------------------------------------

def export_blast_details(blast_dir, target, specificity_result):
    """Write BLAST detail files: HSP summary + per-residue specificity.

    Creates:
      blast/{gene}_blast_hsps.csv         — one row per HSP with call
      blast/{gene}_blast_specificity.csv  — one row per residue with top hit info
    """
    blast_dir = Path(blast_dir)
    blast_dir.mkdir(parents=True, exist_ok=True)
    gene = target.gene_name

    hits = specificity_result.full_blast_hits
    id_scores = specificity_result.residue_identity_scores
    res_spec = specificity_result.residue_specificity
    if not id_scores and not res_spec:
        return

    threshold_pct = SPECIFICITY_IDENTITY_THRESHOLD * 100

    # --- HSP summary table ---------------------------------------------------
    if hits:
        hsps_path = blast_dir / f"{gene}_blast_hsps.csv"
        sorted_hits = sorted(hits, key=lambda h: h.get("identity", 0), reverse=True)
        with open(hsps_path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow([
                "off_target", "accession", "hsp_range", "hsp_length",
                "identical_positions", "identity_pct", "evalue", "call",
            ])
            for h in sorted_hits:
                ident = h.get("identity", 0.0)
                alen = h.get("align_length", 0)
                qs = h.get("query_start", "")
                qe = h.get("query_end", "")
                title = h.get("title", "")
                # Extract short protein name from sp|...|NAME description
                short_name = title
                if title.startswith("sp|"):
                    parts = title.split("|")
                    if len(parts) >= 3:
                        short_name = parts[2].split(" OS=")[0]
                if ident < 0.40 or alen < 30:
                    call = "filtered (<40% or <30aa)"
                elif ident >= SPECIFICITY_IDENTITY_THRESHOLD:
                    call = "non-specific"
                else:
                    call = "specific"
                writer.writerow([
                    short_name, h.get("accession", ""),
                    f"{qs}-{qe}", alen,
                    h.get("identities", ""),
                    f"{ident * 100:.1f}",
                    f"{h.get('evalue', 0):.2e}",
                    call,
                ])
        logger.info("  Wrote %d BLAST HSPs to %s", len(sorted_hits), hsps_path.name)

    # --- Per-residue specificity table ---------------------------------------
    # Build residue → top hit mapping from HSPs
    # For each position, find the highest-identity qualifying HSP that covers it
    residue_top_hit = {}  # {resnum: (identity, accession, title)}
    for h in (hits or []):
        ident = h.get("identity", 0.0)
        alen = h.get("align_length", 0)
        if ident < 0.40 or alen < 30:
            continue
        acc = h.get("accession", "")
        title = h.get("title", "")
        for pos in range(h.get("query_start", 0), h.get("query_end", 0) + 1):
            current = residue_top_hit.get(pos)
            if current is None or ident > current[0]:
                residue_top_hit[pos] = (ident, acc, title)

    out_path = blast_dir / f"{gene}_blast_specificity.csv"
    with open(out_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "residue_num", "aa", "max_off_target_identity_pct",
            "top_hit_accession", "top_hit_protein",
            "specific", "threshold_pct",
        ])
        for i, aa in enumerate(target.sequence, 1):
            max_id = id_scores.get(i, 0.0)
            top = residue_top_hit.get(i)
            acc = top[1] if top else ""
            prot = top[2] if top else ""
            # Trim sp|...|NAME_HUMAN ... to just the gene-level name
            if prot.startswith("sp|"):
                parts = prot.split("|")
                if len(parts) >= 3:
                    prot = parts[2].split(" OS=")[0]  # e.g. "EGFR_HUMAN Epidermal growth factor receptor"
            spec_val = res_spec.get(i)
            if spec_val is True:
                call = "yes"
            elif spec_val is False:
                call = "no"
            else:
                call = "not_assessed"
            writer.writerow([
                i, aa, f"{max_id * 100:.1f}", acc, prot,
                call, f"{threshold_pct:.0f}",
            ])
    logger.info("  Wrote %d residue BLAST details to %s",
                 len(target.sequence), out_path.name)


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------

def export_results_csv(run_dir, targets, epitope_scores, target_metrics):
    """
    Write main results table: one row per epitope patch across all targets.
    """
    csv_path = Path(run_dir) / ".epitope_candidates.csv"
    rows = []

    for target in targets:
        uid = target.uniprot_id
        scores = epitope_scores.get(uid, [])
        metric = target_metrics.get(uid, {})

        if not scores:
            # Write a row indicating no epitope space
            rows.append({
                "target": target.gene_name,
                "uniprot_id": uid,
                "patch_id": "—",
                "rank": "—",
                "composite_score": 0.0,
                "patch_area_A2": 0.0,
                "n_residues": 0,
                "residue_range": "No qualifying epitope space",
                "avg_distance_A": 0.0,
                "cyno_identity": 0.0,
                "max_off_target_identity": 0.0,
                "area_score": 0.0,
                "distance_score": 0.0,
                "conservation_score": 0.0,
                "specificity_score": 0.0,
                "accessibility_score": 0.0,
                "total_epitope_A2": metric.get("total_epitope_area_a2", 0.0),
                "epitope_fraction": metric.get("epitope_fraction", 0.0),
            })
            continue

        for s in scores:
            rows.append({
                "target": target.gene_name,
                "uniprot_id": uid,
                "patch_id": s.patch_id,
                "rank": s.rank,
                "composite_score": round(s.composite_score, 4),
                "patch_area_A2": round(s.patch_area_a2, 1),
                "n_residues": s.n_residues,
                "residue_range": s.residue_range,
                "avg_distance_A": round(s.avg_distance_a, 1),
                "cyno_identity": round(s.cyno_identity, 4),
                "max_off_target_identity": round(s.max_off_target_identity, 4),
                "area_score": round(s.area_score, 4),
                "distance_score": round(s.distance_score, 4),
                "conservation_score": round(s.conservation_score, 4),
                "specificity_score": round(s.specificity_score, 4),
                "accessibility_score": round(s.accessibility_score, 4),
                "total_epitope_A2": round(metric.get("total_epitope_area_a2", 0.0), 1),
                "epitope_fraction": round(metric.get("epitope_fraction", 0.0), 4),
            })

    if rows:
        fieldnames = list(rows[0].keys())
        with open(csv_path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
        logger.info("  Wrote %s (%d rows)", csv_path.name, len(rows))


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------

def export_results_xlsx(run_dir, targets, epitope_scores, target_metrics,
                        membranes, conservation_results):
    """
    Write color-formatted Excel workbook following Cartography style.

    Sheet 1: Summary — one row per target with overall epitope metric
    Sheet 2: Epitope Patches — one row per patch, all scoring components
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    xlsx_path = Path(run_dir) / "epitope_candidates.xlsx"
    wb = Workbook()

    # --- Color fills ---
    header_fill = PatternFill(start_color="28154C", end_color="28154C", fill_type="solid")
    header_font = Font(name="Consolas", bold=True, color="FFFFFF", size=10)
    data_font = Font(name="Consolas", size=10)
    score_font = Font(name="Consolas", size=10, bold=True)
    epitope_fill = PatternFill(start_color="6BC291", end_color="6BC291", fill_type="solid")
    no_space_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # === Sheet 1: Summary ===
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_headers = [
        "Target", "UniProt ID", "Topology", "# TM Segments",
        "# Epitope Patches", "Total Epitope Area (A²)",
        "Ectodomain Area (A²)", "Epitope Fraction",
        "Best Composite Score",
    ]
    for col, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, target in enumerate(targets, 2):
        uid = target.uniprot_id
        scores = epitope_scores.get(uid, [])
        metric = target_metrics.get(uid, {})
        membrane = membranes.get(uid)

        values = [
            target.gene_name,
            uid,
            membrane.topology_type if membrane else "—",
            len(membrane.tm_segments) if membrane else 0,
            len(scores),
            round(metric.get("total_epitope_area_a2", 0.0), 1),
            round(metric.get("total_ectodomain_area_a2", 0.0), 1),
            round(metric.get("epitope_fraction", 0.0), 4),
            round(metric.get("best_score", 0.0), 4),
        ]

        row_fill = epitope_fill if scores else no_space_fill
        for col, val in enumerate(values, 1):
            cell = ws_summary.cell(row=row_idx, column=col, value=val)
            cell.font = data_font
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # Auto-width columns
    for col in range(1, len(summary_headers) + 1):
        ws_summary.column_dimensions[get_column_letter(col)].width = 18

    # === Sheet 2: Epitope Patches ===
    ws_patches = wb.create_sheet("Epitope Patches")

    patch_headers = [
        "Target", "UniProt ID", "Patch ID", "Rank",
        "Composite Score", "Area (A²)", "# Residues", "Residue Range",
        "Avg Distance (A)", "Cyno Identity", "Max Off-Target ID",
        "Area Score", "Distance Score", "Conservation Score",
        "Specificity Score", "Accessibility Score",
    ]
    for col, h in enumerate(patch_headers, 1):
        cell = ws_patches.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row_idx = 2
    for target in targets:
        uid = target.uniprot_id
        scores = epitope_scores.get(uid, [])

        if not scores:
            cell = ws_patches.cell(row=row_idx, column=1, value=target.gene_name)
            cell.font = data_font
            cell.fill = no_space_fill
            cell = ws_patches.cell(row=row_idx, column=2, value=uid)
            cell.font = data_font
            cell.fill = no_space_fill
            cell = ws_patches.cell(row=row_idx, column=8, value="No qualifying epitope space")
            cell.font = data_font
            cell.fill = no_space_fill
            row_idx += 1
            continue

        for s in scores:
            values = [
                target.gene_name, uid, s.patch_id, s.rank,
                round(s.composite_score, 4),
                round(s.patch_area_a2, 1),
                s.n_residues, s.residue_range,
                round(s.avg_distance_a, 1),
                round(s.cyno_identity, 4),
                round(s.max_off_target_identity, 4),
                round(s.area_score, 4),
                round(s.distance_score, 4),
                round(s.conservation_score, 4),
                round(s.specificity_score, 4),
                round(s.accessibility_score, 4),
            ]
            for col, val in enumerate(values, 1):
                cell = ws_patches.cell(row=row_idx, column=col, value=val)
                cell.font = score_font if col == 5 else data_font
                cell.fill = epitope_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
            row_idx += 1

    for col in range(1, len(patch_headers) + 1):
        ws_patches.column_dimensions[get_column_letter(col)].width = 16

    wb.save(str(xlsx_path))
    logger.info("  Wrote %s", xlsx_path.name)


# ---------------------------------------------------------------------------
# Annotated PDB export
# ---------------------------------------------------------------------------

def export_annotated_pdb(run_dir, target, structure, membrane,
                         spatial_filter, surface_analysis,
                         conservation_result, scores,
                         pdb_subdir="Structures"):
    """
    Write a richly annotated PDB with per-criteria B-factor tiers.

    B-factor encodes which pipeline criteria each residue passes, enabling
    PyMOL/ChimeraX selection by filter stage:

        B-factor   Tier                               PyMOL selection
        --------   ----                               ---------------
         -20       Non-target chain (Ab, etc.)        select other, b < -10
           0       Intracellular                      select ic, b > -5 and b < 5
          10       Transmembrane                      select tm, b > 5 and b < 15
          25       Extracellular (below dist cutoff)  select ec_near, b > 20 and b < 35
          40       Distance-qualified (>=100A)        select dist_pass, b > 35 and b < 45
          55       + Surface exposed (SASA > 25%)     select surf_pass, b > 50 and b < 65
          70       + Cyno conserved (>=98%)           select cons_pass, b > 65 and b < 78
          80-95    EPITOPE PATCH (composite * 100)    select epitope, b > 78
          99       Membrane reference (3 proximal)    select mem_ref, b > 97

    Occupancy column stores the distance from membrane (Angstroms),
    enabling continuous coloring: `spectrum q, blue_white_red, chain C`

    Args:
        run_dir: Run output directory.
        target: TargetInfo.
        structure: StructureResult.
        membrane: MembraneAnnotation.
        spatial_filter: SpatialFilter (or None).
        surface_analysis: SurfaceAnalysis (or None).
        conservation_result: ConservationResult (or None).
        scores: List of EpitopeScore objects.
    """
    import numpy as np
    parser = PDBParser(QUIET=True)
    bio_structure = parser.get_structure(target.gene_name, structure.pdb_path)

    # --- Build lookup sets for each criterion ---

    # Epitope patch residues and their scores
    patch_scores = {}
    for s in scores:
        for resnum in s.patch.residue_numbers:
            patch_scores[resnum] = s.composite_score * 100

    # Distance-qualified residues
    dist_qualified = set()
    if spatial_filter:
        dist_qualified = set(spatial_filter.qualifying_residues)

    # Surface-exposed residues
    surf_exposed = set()
    if surface_analysis:
        surf_exposed = set(surface_analysis.exposed_residues)

    # Cyno-conserved residues
    cyno_conserved = set()
    if conservation_result:
        for resnum, is_conserved in conservation_result.residue_conservation.items():
            if is_conserved is True:
                cyno_conserved.add(resnum)

    # Distances from membrane (for occupancy column)
    residue_distances = {}
    if spatial_filter:
        residue_distances = spatial_filter.residue_distances

    # --- Find 3 membrane-proximal reference residues ---
    # These are the 3 Cα atoms on the target chain with smallest positive
    # distance along the membrane normal — the measurement origin
    membrane_ref_residues = set()
    target_chain = None
    for model in bio_structure:
        if structure.chain_id in model:
            target_chain = model[structure.chain_id]
        else:
            chains = list(model.get_chains())
            target_chain = chains[0] if chains else None
        break

    if target_chain is not None:
        proximal_candidates = []
        for res in target_chain:
            res_id = res.get_id()
            if res_id[0] != " ":
                continue
            resnum = res_id[1]
            if "CA" in res:
                # Only consider extracellular residues for membrane reference
                topo = membrane.residue_topology.get(resnum)
                if topo != "extracellular":
                    continue
                ca_pos = res["CA"].get_vector().get_array()
                d = np.dot(
                    ca_pos - membrane.membrane_center,
                    membrane.membrane_normal,
                )
                if d > 0:
                    proximal_candidates.append((d, resnum))

        # Sort by distance ascending — smallest positive = closest to
        # membrane plane on the extracellular side
        proximal_candidates.sort(key=lambda x: x[0])
        for d, resnum in proximal_candidates[:3]:
            membrane_ref_residues.add(resnum)

    # --- Compute ecto-axis rotation for single-pass TM ---
    # Rotate PDB coordinates so anchor→farthest aligns with Y-up and
    # anchor sits at the origin. This makes kinked ectodomains (e.g. ROR1)
    # extend straight up from the bilayer in PyMOL.
    _rotation_mat = None
    _rotation_origin = None
    if (membrane and spatial_filter
            and getattr(spatial_filter, 'anchor_resnum', None) is not None
            and getattr(spatial_filter, 'farthest_resnum', None) is not None):
        from epitope_pipeline.utils import extract_ca_coords as _extract_ca
        _ca = _extract_ca(structure.pdb_path, structure.chain_id)
        _a = _ca.get(spatial_filter.anchor_resnum)
        _f = _ca.get(spatial_filter.farthest_resnum)
        if _a is not None and _f is not None:
            _axis = _f - _a
            if np.linalg.norm(_axis) > 1.0:
                _rotation_mat = _rotation_matrix_align(
                    _axis / np.linalg.norm(_axis),
                    np.array([0.0, 1.0, 0.0]))
                _rotation_origin = _a.copy()
                logger.info(
                    "  %s: Rotating PDB coordinates "
                    "(ecto-axis -> Y-up, anchor res %d at origin)",
                    target.gene_name, spatial_filter.anchor_resnum)

    # --- Assign B-factors and occupancy ---
    for model in bio_structure:
        for chain in model:
            chain_id = chain.get_id()
            is_target_chain = (chain_id == structure.chain_id)

            for residue in chain:
                res_id = residue.get_id()
                if res_id[0] != " ":
                    continue
                resnum = res_id[1]

                if not is_target_chain:
                    bfactor = -20.0
                    occupancy = 0.0
                elif resnum in membrane_ref_residues:
                    bfactor = 99.0
                    occupancy = residue_distances.get(resnum, 0.0)
                elif resnum in patch_scores:
                    bfactor = max(80.0, min(95.0, patch_scores[resnum]))
                    occupancy = residue_distances.get(resnum, 0.0)
                elif resnum in cyno_conserved and resnum in surf_exposed:
                    bfactor = 70.0
                    occupancy = residue_distances.get(resnum, 0.0)
                elif resnum in surf_exposed:
                    bfactor = 55.0
                    occupancy = residue_distances.get(resnum, 0.0)
                elif resnum in dist_qualified:
                    bfactor = 40.0
                    occupancy = residue_distances.get(resnum, 0.0)
                else:
                    topo = membrane.residue_topology.get(resnum, "")
                    if topo == "extracellular":
                        bfactor = 25.0
                        occupancy = residue_distances.get(resnum, 0.0)
                    elif topo == "transmembrane":
                        bfactor = 10.0
                        occupancy = 0.0
                    elif topo == "intracellular":
                        bfactor = 0.0
                        occupancy = 0.0
                    else:
                        bfactor = -20.0
                        occupancy = 0.0

                for atom in residue:
                    atom.set_bfactor(bfactor)
                    atom.set_occupancy(occupancy)
                    if _rotation_mat is not None:
                        pos = atom.get_vector().get_array()
                        atom.set_coord(
                            _rotation_mat @ (pos - _rotation_origin))

    # --- Write PDB (dot-prefixed to hide from file browsers) ---
    out_path = Path(run_dir) / pdb_subdir / ".{}_epitope.pdb".format(
        target.gene_name.lower()
    )
    io = PDBIO()
    io.set_structure(bio_structure)
    io.save(str(out_path))

    # --- Prepend REMARK records with scheme + PyMOL commands ---
    with open(out_path, "r") as f:
        content = f.read()

    # Summary stats for the header
    n_dist = len(dist_qualified)
    n_surf = len(surf_exposed)
    n_cons = len(cyno_conserved & surf_exposed)
    n_patch = len(patch_scores)
    ref_resnums = sorted(membrane_ref_residues)

    remarks = [
        "REMARK 999",
        "REMARK 999 EPITOPE PIPELINE — ANNOTATED PDB",
        "REMARK 999 Target: {} ({})".format(target.gene_name, target.uniprot_id),
        "REMARK 999 Structure: {} (chain {}, {})".format(
            structure.pdb_id, structure.chain_id, structure.method),
        "REMARK 999 Distance threshold: {:.0f} A from membrane".format(
            spatial_filter.min_distance_threshold if spatial_filter else 0),
        "REMARK 999",
        "REMARK 999 ---- B-FACTOR ENCODING (per-criteria tiers) ----",
        "REMARK 999",
        "REMARK 999   B = -20  Non-target chain (antibody, etc.)",
        "REMARK 999   B =   0  Intracellular",
        "REMARK 999   B =  10  Transmembrane",
        "REMARK 999   B =  25  Extracellular (below distance cutoff)",
        "REMARK 999   B =  40  Distance qualified (>= {:.0f} A)        [{} residues]".format(
            spatial_filter.min_distance_threshold if spatial_filter else 0, n_dist),
        "REMARK 999   B =  55  + Surface exposed (rel SASA > 25%)    [{} residues]".format(n_surf),
        "REMARK 999   B =  70  + Cyno conserved (>= 98% identity)    [{} residues]".format(n_cons),
        "REMARK 999   B = 80-95 EPITOPE PATCH (composite score x100) [{} residues]".format(n_patch),
        "REMARK 999   B =  99  Membrane reference (3 most proximal)  [res {}]".format(
            ", ".join(str(r) for r in ref_resnums)),
        "REMARK 999",
        "REMARK 999 ---- OCCUPANCY COLUMN ----",
        "REMARK 999   Occupancy = distance from membrane plane (Angstroms)",
        "REMARK 999   (target chain extracellular residues only)",
        "REMARK 999",
        "REMARK 999 ---- PYMOL QUICK SELECTIONS ----",
        "REMARK 999",
        "REMARK 999   # Load and orient",
        "REMARK 999   load .{}_epitope.pdb".format(target.gene_name.lower()),
        "REMARK 999",
        "REMARK 999   # Select by criteria tier",
        "REMARK 999   select target_chain, chain {}".format(structure.chain_id),
        "REMARK 999   select other_chains, b < -10",
        "REMARK 999   select intracellular, chain {} and b > -5 and b < 5".format(structure.chain_id),
        "REMARK 999   select transmembrane, chain {} and b > 5 and b < 15".format(structure.chain_id),
        "REMARK 999   select ec_near_membrane, chain {} and b > 20 and b < 35".format(structure.chain_id),
        "REMARK 999   select dist_qualified, chain {} and b > 35".format(structure.chain_id),
        "REMARK 999   select surface_exposed, chain {} and b > 50".format(structure.chain_id),
        "REMARK 999   select cyno_conserved, chain {} and b > 65 and b < 78".format(structure.chain_id),
        "REMARK 999   select epitope_patch, chain {} and b > 78 and b < 97".format(structure.chain_id),
        "REMARK 999   select membrane_ref, chain {} and b > 97".format(structure.chain_id),
        "REMARK 999",
        "REMARK 999   # Recommended visualization",
        "REMARK 999   hide everything",
        "REMARK 999   show cartoon, target_chain",
        "REMARK 999   spectrum b, white_blue_red, chain {}, minimum=-20, maximum=99".format(
            structure.chain_id),
        "REMARK 999   show spheres, membrane_ref",
        "REMARK 999   color magenta, membrane_ref",
        "REMARK 999   show surface, epitope_patch",
        "REMARK 999   color green, epitope_patch",
        "REMARK 999",
    ]

    # Add per-patch details
    for s in scores:
        resnums = sorted(s.patch.residue_numbers)
        remarks.append(
            "REMARK 999 Patch {}: score={:.3f}, {} residues, {:.0f} A2, "
            "cyno={:.0f}%".format(
                s.patch_id, s.composite_score, len(resnums),
                s.patch.total_sasa_a2, s.cyno_identity * 100,
            )
        )
        remarks.append(
            "REMARK 999   Residues: {}".format(s.residue_range)
        )

    if ref_resnums:
        remarks.append("REMARK 999")
        remarks.append(
            "REMARK 999 Membrane reference residues (measurement origin):"
        )
        for r in ref_resnums:
            d = residue_distances.get(r, 0.0)
            remarks.append(
                "REMARK 999   Residue {}: {:.1f} A from membrane".format(r, d)
            )

    remarks.append("REMARK 999")

    with open(out_path, "w") as f:
        f.write("\n".join(remarks) + "\n" + content)

    logger.info("  Wrote annotated PDB: %s", out_path.name)

    # --- Write companion .pml script with clickable selections ---
    _write_pymol_script(
        run_dir, target, structure, membrane, spatial_filter,
        dist_qualified, surf_exposed, cyno_conserved,
        patch_scores, membrane_ref_residues, residue_distances,
        scores, pdb_subdir=pdb_subdir,
        pdb_rotated=(_rotation_mat is not None),
    )


# ---------------------------------------------------------------------------
# Membrane bilayer CGO visualization
# ---------------------------------------------------------------------------

def generate_membrane_cgo_pml(membrane, pdb_path, chain_id,
                               object_prefix="", translate_vec=None,
                               normal_override=None, center_override=None):
    """
    Generate PML lines for CGO membrane bilayer visualization.

    Creates a lipid-bilayer-like rendering using scattered CGO spheres:
      - Headgroup bands at the extracellular/intracellular surfaces
        (mixed red, orange, blue, gray — mimicking phospholipid headgroups)
      - Gray tail spheres filling the hydrophobic core
      - Exclusion zone around the TM helix so protein threads through

    Args:
        membrane: MembraneAnnotation with center, normal, half_thickness.
        pdb_path: Path to PDB file for computing disk radius.
        chain_id: Chain ID in the PDB.
        object_prefix: Prefix for PyMOL object names (e.g., "Distal_").
        translate_vec: 3D translation vector for bispecific side-by-side view.
        normal_override: Optional replacement for membrane normal (e.g., ecto-axis).
        center_override: Optional replacement for membrane center.

    Returns:
        List of PML lines to insert into a .pml script.
    """
    from epitope_pipeline.utils import extract_ca_coords

    center = np.asarray(center_override, dtype=float) if center_override is not None else membrane.membrane_center
    normal = np.asarray(normal_override, dtype=float) if normal_override is not None else membrane.membrane_normal
    ht = membrane.membrane_half_thickness

    # In-plane basis vectors for the membrane plane
    ref = np.array([1.0, 0.0, 0.0]) if abs(normal[0]) < 0.9 else np.array([0.0, 1.0, 0.0])
    v1 = np.cross(normal, ref)
    v1 = v1 / np.linalg.norm(v1)
    v2 = np.cross(normal, v1)

    # Compute disk radius from structure extent projected onto membrane plane
    ca_coords = extract_ca_coords(pdb_path, chain_id)
    max_radial = 0.0
    for ca_pos in ca_coords.values():
        v = ca_pos - center
        v_in_plane = v - np.dot(v, normal) * normal
        radial_dist = np.linalg.norm(v_in_plane)
        if radial_dist > max_radial:
            max_radial = radial_dist
    radius = max_radial + 15.0

    # TM exclusion zone: project TM residues onto membrane plane
    tm_plane_pos = []
    for resnum, topo in membrane.residue_topology.items():
        if topo == "transmembrane" and resnum in ca_coords:
            v = ca_coords[resnum] - center
            tm_plane_pos.append((np.dot(v, v1), np.dot(v, v2)))
    exclusion_r2 = 10.0 ** 2  # 10A exclusion radius around TM

    # Generate random lipid positions on the membrane plane (fixed seed)
    rng = np.random.RandomState(42)
    n_target = 350
    lipid_positions = []  # (in-plane x, in-plane y)
    attempts = 0
    while len(lipid_positions) < n_target and attempts < n_target * 5:
        attempts += 1
        r = radius * np.sqrt(rng.random())
        theta = rng.random() * 2 * np.pi
        px, py = r * np.cos(theta), r * np.sin(theta)
        excluded = any((px - tx) ** 2 + (py - ty) ** 2 < exclusion_r2
                        for tx, ty in tm_plane_pos)
        if not excluded:
            lipid_positions.append((px, py))

    # Headgroup atom colors — mostly gray with sparse red/orange accents
    head_colors = [
        [0.75, 0.75, 0.75],  # gray
        [0.75, 0.75, 0.75],  # gray
        [0.75, 0.75, 0.75],  # gray
        [0.75, 0.75, 0.75],  # gray
        [0.80, 0.25, 0.15],  # red (sparse)
        [0.85, 0.55, 0.15],  # orange (sparse)
    ]
    tail_color = [0.65, 0.65, 0.65]
    head_sphere_r = 2.0
    tail_sphere_r = 1.7

    # Build sphere data: (3D position, color, radius, layer_type)
    spheres_head = []
    spheres_tail = []

    for px, py in lipid_positions:
        # 3D base position on the membrane plane
        base = center + v1 * px + v2 * py
        if translate_vec is not None:
            base = base.copy() + translate_vec

        # Random color for this lipid's headgroups
        hc_idx = rng.randint(0, len(head_colors))
        hcolor = head_colors[hc_idx]

        # Top headgroup (extracellular surface) — 1-2 spheres
        h_offset = ht + rng.uniform(-1.0, 1.0)
        pos = base + normal * h_offset
        spheres_head.append((pos, hcolor, head_sphere_r))
        # Second headgroup atom slightly offset
        h2_offset = h_offset - rng.uniform(1.5, 3.0)
        dx = rng.uniform(-1.5, 1.5)
        dy = rng.uniform(-1.5, 1.5)
        pos2 = base + normal * h2_offset + v1 * dx + v2 * dy
        hc2 = head_colors[rng.randint(0, len(head_colors))]
        spheres_head.append((pos2, hc2, head_sphere_r * 0.85))

        # Bottom headgroup (intracellular surface) — 1-2 spheres
        h_offset_b = -ht + rng.uniform(-1.0, 1.0)
        pos_b = base + normal * h_offset_b
        spheres_head.append((pos_b, hcolor, head_sphere_r))
        h2_offset_b = h_offset_b + rng.uniform(1.5, 3.0)
        pos_b2 = base + normal * h2_offset_b + v1 * dx + v2 * dy
        spheres_head.append((pos_b2, hc2, head_sphere_r * 0.85))

        # Tail spheres: zigzag chain filling the core for each leaflet
        # Upper leaflet tail (from top headgroup inward)
        n_tail_upper = rng.randint(3, 5)
        tail_dx = rng.uniform(-1.0, 1.0)
        tail_dy = rng.uniform(-1.0, 1.0)
        for j in range(n_tail_upper):
            t = (j + 1) / (n_tail_upper + 1)
            t_offset = (ht - 3.0) * (1.0 - t)  # from near headgroup toward center
            jitter_d = rng.uniform(-1.0, 1.0)
            jitter_v1 = tail_dx * ((-1) ** j) + rng.uniform(-0.5, 0.5)
            jitter_v2 = tail_dy * ((-1) ** j) + rng.uniform(-0.5, 0.5)
            t_pos = base + normal * t_offset + v1 * jitter_v1 + v2 * jitter_v2
            shade = 0.60 + rng.uniform(0, 0.12)
            spheres_tail.append((t_pos, [shade, shade, shade], tail_sphere_r))

        # Lower leaflet tail (from bottom headgroup inward)
        n_tail_lower = rng.randint(3, 5)
        for j in range(n_tail_lower):
            t = (j + 1) / (n_tail_lower + 1)
            t_offset = -(ht - 3.0) * (1.0 - t)
            jitter_v1 = tail_dx * ((-1) ** j) + rng.uniform(-0.5, 0.5)
            jitter_v2 = tail_dy * ((-1) ** j) + rng.uniform(-0.5, 0.5)
            t_pos = base + normal * t_offset + v1 * jitter_v1 + v2 * jitter_v2
            shade = 0.60 + rng.uniform(0, 0.12)
            spheres_tail.append((t_pos, [shade, shade, shade], tail_sphere_r))

    # Object names
    head_name = "{}Membrane_Heads".format(object_prefix)
    tail_name = "{}Membrane_Tails".format(object_prefix)
    group_name = "{}Membrane".format(object_prefix)

    # Generate PML with CGO SPHERE primitives
    lines = [
        "# --- Membrane bilayer (lipid sphere model) ---",
        "python",
        "from pymol.cgo import *",
        "from pymol import cmd",
        "",
    ]

    # Headgroup CGO
    lines.append("heads = [")
    for pos, color, r in spheres_head:
        lines.append("    COLOR, {:.2f}, {:.2f}, {:.2f}, SPHERE, {:.2f}, {:.2f}, {:.2f}, {:.2f},".format(
            color[0], color[1], color[2], pos[0], pos[1], pos[2], r,
        ))
    lines.append("]")
    lines.append("cmd.load_cgo(heads, '{}')".format(head_name))
    lines.append("")

    # Tail CGO
    lines.append("tails = [")
    for pos, color, r in spheres_tail:
        lines.append("    COLOR, {:.2f}, {:.2f}, {:.2f}, SPHERE, {:.2f}, {:.2f}, {:.2f}, {:.2f},".format(
            color[0], color[1], color[2], pos[0], pos[1], pos[2], r,
        ))
    lines.append("]")
    lines.append("cmd.load_cgo(tails, '{}')".format(tail_name))
    lines.append("")

    lines.append("cmd.group('{}', '{} {}')".format(group_name, head_name, tail_name))
    lines.append("python end")
    lines.append("")

    return lines


def generate_shared_membrane_cgo_pml(membrane_a, pdb_path_a, chain_a,
                                      membrane_b, pdb_path_b, chain_b,
                                      translate_b=None,
                                      normal_override=None, center_override=None):
    """
    Generate PML lines for a single shared membrane CGO spanning two structures.

    Structure A sits at origin; structure B is translated by translate_b.
    The membrane plane is averaged from both targets' membrane annotations so
    both proteins appear embedded in one continuous lipid bilayer.

    Args:
        membrane_a: MembraneAnnotation for the first (distal) target.
        pdb_path_a: Path to first target's PDB file.
        chain_a: Chain ID for first target.
        membrane_b: MembraneAnnotation for the second (proximal) target.
        pdb_path_b: Path to second target's PDB file.
        chain_b: Chain ID for second target.
        translate_b: 3D translation vector applied to structure B.
        normal_override: Optional replacement for the averaged membrane normal.
        center_override: Optional replacement for the averaged membrane center.

    Returns:
        List of PML lines to insert into a .pml script.
    """
    from epitope_pipeline.utils import extract_ca_coords

    center_a = np.array(membrane_a.membrane_center, dtype=float)
    normal_a = np.array(membrane_a.membrane_normal, dtype=float)
    ht_a = membrane_a.membrane_half_thickness

    center_b = np.array(membrane_b.membrane_center, dtype=float).copy()
    if translate_b is not None:
        center_b += np.asarray(translate_b, dtype=float)
    normal_b = np.array(membrane_b.membrane_normal, dtype=float)
    ht_b = membrane_b.membrane_half_thickness

    # Shared membrane plane: average center, normal, half-thickness
    # (overrides take priority when provided)
    if normal_override is not None:
        normal = np.asarray(normal_override, dtype=float)
    else:
        avg_normal = (normal_a + normal_b) / 2.0
        norm_len = np.linalg.norm(avg_normal)
        normal = avg_normal / norm_len if norm_len > 1e-6 else normal_a
    if center_override is not None:
        center = np.asarray(center_override, dtype=float)
    else:
        center = (center_a + center_b) / 2.0
    ht = (ht_a + ht_b) / 2.0

    # In-plane basis vectors
    ref = np.array([1.0, 0.0, 0.0]) if abs(normal[0]) < 0.9 else np.array([0.0, 1.0, 0.0])
    v1 = np.cross(normal, ref)
    v1 = v1 / np.linalg.norm(v1)
    v2 = np.cross(normal, v1)

    # Compute disk radius from BOTH structures' extents on the membrane plane
    ca_a = extract_ca_coords(pdb_path_a, chain_a)
    ca_b = extract_ca_coords(pdb_path_b, chain_b)

    max_radial = 0.0
    for ca_pos in ca_a.values():
        v = np.array(ca_pos, dtype=float) - center
        v_in_plane = v - np.dot(v, normal) * normal
        radial_dist = np.linalg.norm(v_in_plane)
        if radial_dist > max_radial:
            max_radial = radial_dist

    _tvec_b = np.asarray(translate_b, dtype=float) if translate_b is not None else np.zeros(3)
    for ca_pos in ca_b.values():
        pos = np.array(ca_pos, dtype=float) + _tvec_b
        v = pos - center
        v_in_plane = v - np.dot(v, normal) * normal
        radial_dist = np.linalg.norm(v_in_plane)
        if radial_dist > max_radial:
            max_radial = radial_dist

    radius = max_radial + 15.0

    # TM exclusion zones from BOTH structures
    tm_plane_pos = []
    for resnum, topo in membrane_a.residue_topology.items():
        if topo == "transmembrane" and resnum in ca_a:
            v = np.array(ca_a[resnum], dtype=float) - center
            tm_plane_pos.append((np.dot(v, v1), np.dot(v, v2)))

    for resnum, topo in membrane_b.residue_topology.items():
        if topo == "transmembrane" and resnum in ca_b:
            pos = np.array(ca_b[resnum], dtype=float) + _tvec_b
            v = pos - center
            tm_plane_pos.append((np.dot(v, v1), np.dot(v, v2)))

    exclusion_r2 = 10.0 ** 2

    # Scale lipid count to membrane area (base: 350 for single-target ~50A radius)
    area_ratio = (radius / 50.0) ** 2
    n_target = max(350, min(1200, int(350 * area_ratio)))

    rng = np.random.RandomState(42)
    lipid_positions = []
    attempts = 0
    while len(lipid_positions) < n_target and attempts < n_target * 5:
        attempts += 1
        r = radius * np.sqrt(rng.random())
        theta = rng.random() * 2 * np.pi
        px, py = r * np.cos(theta), r * np.sin(theta)
        excluded = any((px - tx) ** 2 + (py - ty) ** 2 < exclusion_r2
                       for tx, ty in tm_plane_pos)
        if not excluded:
            lipid_positions.append((px, py))

    # Headgroup colors and sizes (same as single-target)
    head_colors = [
        [0.75, 0.75, 0.75],
        [0.75, 0.75, 0.75],
        [0.75, 0.75, 0.75],
        [0.75, 0.75, 0.75],
        [0.80, 0.25, 0.15],
        [0.85, 0.55, 0.15],
    ]
    head_sphere_r = 2.0
    tail_sphere_r = 1.7

    spheres_head = []
    spheres_tail = []

    for px, py in lipid_positions:
        base = center + v1 * px + v2 * py

        hc_idx = rng.randint(0, len(head_colors))
        hcolor = head_colors[hc_idx]

        # Top headgroup
        h_offset = ht + rng.uniform(-1.0, 1.0)
        pos = base + normal * h_offset
        spheres_head.append((pos, hcolor, head_sphere_r))
        h2_offset = h_offset - rng.uniform(1.5, 3.0)
        dx = rng.uniform(-1.5, 1.5)
        dy = rng.uniform(-1.5, 1.5)
        pos2 = base + normal * h2_offset + v1 * dx + v2 * dy
        hc2 = head_colors[rng.randint(0, len(head_colors))]
        spheres_head.append((pos2, hc2, head_sphere_r * 0.85))

        # Bottom headgroup
        h_offset_b = -ht + rng.uniform(-1.0, 1.0)
        pos_b = base + normal * h_offset_b
        spheres_head.append((pos_b, hcolor, head_sphere_r))
        h2_offset_b = h_offset_b + rng.uniform(1.5, 3.0)
        pos_b2 = base + normal * h2_offset_b + v1 * dx + v2 * dy
        spheres_head.append((pos_b2, hc2, head_sphere_r * 0.85))

        # Tail spheres — upper leaflet
        n_tail_upper = rng.randint(3, 5)
        tail_dx = rng.uniform(-1.0, 1.0)
        tail_dy = rng.uniform(-1.0, 1.0)
        for j in range(n_tail_upper):
            t = (j + 1) / (n_tail_upper + 1)
            t_offset = (ht - 3.0) * (1.0 - t)
            jitter_v1 = tail_dx * ((-1) ** j) + rng.uniform(-0.5, 0.5)
            jitter_v2 = tail_dy * ((-1) ** j) + rng.uniform(-0.5, 0.5)
            t_pos = base + normal * t_offset + v1 * jitter_v1 + v2 * jitter_v2
            shade = 0.60 + rng.uniform(0, 0.12)
            spheres_tail.append((t_pos, [shade, shade, shade], tail_sphere_r))

        # Tail spheres — lower leaflet
        n_tail_lower = rng.randint(3, 5)
        for j in range(n_tail_lower):
            t = (j + 1) / (n_tail_lower + 1)
            t_offset = -(ht - 3.0) * (1.0 - t)
            jitter_v1 = tail_dx * ((-1) ** j) + rng.uniform(-0.5, 0.5)
            jitter_v2 = tail_dy * ((-1) ** j) + rng.uniform(-0.5, 0.5)
            t_pos = base + normal * t_offset + v1 * jitter_v1 + v2 * jitter_v2
            shade = 0.60 + rng.uniform(0, 0.12)
            spheres_tail.append((t_pos, [shade, shade, shade], tail_sphere_r))

    head_name = "Membrane_Heads"
    tail_name = "Membrane_Tails"
    group_name = "Membrane"

    lines = [
        "# --- Shared membrane bilayer (lipid sphere model) ---",
        "python",
        "from pymol.cgo import *",
        "from pymol import cmd",
        "",
    ]

    lines.append("heads = [")
    for pos, color, r in spheres_head:
        lines.append("    COLOR, {:.2f}, {:.2f}, {:.2f}, SPHERE, {:.2f}, {:.2f}, {:.2f}, {:.2f},".format(
            color[0], color[1], color[2], pos[0], pos[1], pos[2], r,
        ))
    lines.append("]")
    lines.append("cmd.load_cgo(heads, '{}')".format(head_name))
    lines.append("")

    lines.append("tails = [")
    for pos, color, r in spheres_tail:
        lines.append("    COLOR, {:.2f}, {:.2f}, {:.2f}, SPHERE, {:.2f}, {:.2f}, {:.2f}, {:.2f},".format(
            color[0], color[1], color[2], pos[0], pos[1], pos[2], r,
        ))
    lines.append("]")
    lines.append("cmd.load_cgo(tails, '{}')".format(tail_name))
    lines.append("")

    lines.append("cmd.group('{}', '{} {}')".format(group_name, head_name, tail_name))
    lines.append("python end")
    lines.append("")

    return lines


def _write_pymol_script(run_dir, target, structure, membrane, spatial_filter,
                         dist_qualified, surf_exposed, cyno_conserved,
                         patch_scores, membrane_ref_residues,
                         residue_distances, scores,
                         pdb_subdir="Structures",
                         pdb_rotated=False):
    """
    Write a .pml PyMOL script that creates named, clickable selections
    in the PyMOL object panel (right sidebar).

    Selections use explicit residue numbers, not B-factor ranges,
    so they're exact and robust.

    Colors use the Cartography palette:
        #D3D3D3  light gray       — non-target chains
        #E0F3DB  pale green       — ectodomain proximal to membrane
        #2E95D2  blue             — distance-qualified but buried
        #18B5CB  teal             — surface exposed
        #6BC291  green            — epitope patches (most visible)
        #28154C  dark purple      — cyno conserved (not in patch)
    """
    gene = target.gene_name.lower()
    ch = structure.chain_id
    pdb_name = "{}_epitope".format(gene)
    pml_path = Path(run_dir) / pdb_subdir / "{}_epitope.pml".format(gene)
    dist_thresh = spatial_filter.min_distance_threshold if spatial_filter else 0

    # Helper: format a set of residue numbers into a PyMOL resi selection string
    # e.g. {1,2,3,5,6,10} -> "1-3+5-6+10"
    def resi_str(resnums):
        if not resnums:
            return "none"
        nums = sorted(resnums)
        ranges = []
        start = nums[0]
        end = nums[0]
        for n in nums[1:]:
            if n == end + 1:
                end = n
            else:
                ranges.append("{}-{}".format(start, end) if end > start else str(start))
                start = end = n
        ranges.append("{}-{}".format(start, end) if end > start else str(start))
        return "+".join(ranges)

    # Build residue sets for each tier (exclusive — each residue in exactly one)
    patch_set = set(patch_scores.keys())
    ref_set = set(membrane_ref_residues)
    conserved_surf = (cyno_conserved & surf_exposed) - patch_set - ref_set
    surf_only = surf_exposed - cyno_conserved - patch_set - ref_set
    dist_only = dist_qualified - surf_exposed - patch_set - ref_set

    # Extracellular below distance threshold
    all_ec = set()
    for resnum, topo in membrane.residue_topology.items():
        if topo == "extracellular":
            all_ec.add(resnum)
    ec_near = all_ec - dist_qualified - patch_set - ref_set

    # Cumulative sets (for "show me everything passing X and above")
    all_dist_pass = dist_qualified | patch_set  # everything >= distance threshold
    all_surf_pass = (surf_exposed & dist_qualified) | patch_set
    all_cons_pass = (cyno_conserved & surf_exposed & dist_qualified) | patch_set

    # --- Cartography palette as PyMOL RGB [0-1] ---
    # #D3D3D3 → light gray,  #E0F3DB → pale green,  #6BC291 → green
    # #18B5CB → teal,  #2E95D2 → blue,  #28154C → dark purple
    lines = [
        "# Epitope Pipeline — PyMOL visualization script",
        "# Target: {} ({})".format(target.gene_name, target.uniprot_id),
        "# Structure: {} chain {}".format(structure.pdb_id, ch),
        "# Usage: Open this .pml in PyMOL, or run: @{}_epitope.pml".format(gene),
        "#",
        "# All selections appear in the right panel as clickable objects.",
        "# Colors: Cartography palette",
        "",
        "# --- Load structure ---",
        "load .{}_epitope.pdb, {}".format(gene, pdb_name),
        "",
        "# --- Cartography palette ---",
        "set_color carto_gray,       [0.827, 0.827, 0.827]",
        "set_color carto_palegreen,   [0.878, 0.953, 0.859]",
        "set_color carto_green,       [0.420, 0.761, 0.569]",
        "set_color carto_teal,        [0.094, 0.710, 0.796]",
        "set_color carto_blue,        [0.180, 0.584, 0.824]",
        "set_color carto_purple,      [0.157, 0.082, 0.298]",
        "",
        "# --- Base display ---",
        "hide everything",
        "show cartoon, {} and chain {}".format(pdb_name, ch),
        "color white, {} and chain {}".format(pdb_name, ch),
        "set cartoon_transparency, 0.3, {} and chain {}".format(pdb_name, ch),
        "",
    ]

    # For AlphaFold full-length structures, hide TM + cytoplasmic domains
    # to keep the view focused on the ectodomain.
    # For multi-pass proteins, show everything (TM bundle is integral to the structure).
    # For single-pass / GPI, hide only residues below the membrane plane using
    # spatial distance data (more reliable than topology classification alone).
    # No topology coloring — white cartoon + green patches only

    # =============================================
    # EPITOPE PATCHES — grouped under "Epitopes"
    # =============================================
    patch_names = []
    for s in scores:
        resnums = set(s.patch.residue_numbers) - ref_set
        sel = "Patch_{}".format(s.patch_id)
        patch_names.append(sel)
        lines.append("# PATCH {} — score={:.3f}, {} residues, {:.0f} A2, cyno={:.0f}%".format(
            s.patch_id, s.composite_score, len(s.patch.residue_numbers),
            s.patch.total_sasa_a2, s.cyno_identity * 100))
        lines.append("select {}, {} and chain {} and resi {}".format(
            sel, pdb_name, ch, resi_str(resnums)))
        lines.append("color carto_green, {}".format(sel))
        lines.append("show surface, {}".format(sel))
        lines.append("set transparency, 0.3, {}".format(sel))
        lines.append("set cartoon_transparency, 0.0, {}".format(sel))
        lines.append("")

    if patch_set:
        all_patch = patch_set - ref_set
        lines.append("select Combined, {} and chain {} and resi {}".format(
            pdb_name, ch, resi_str(all_patch)))
        patch_names.append("Combined")
        lines.append("")

    if patch_names:
        lines.append("group Epitopes, {}".format(" ".join(patch_names)))
        lines.append("group Epitopes, close")
        lines.append("")

    # =============================================
    # FILTER TIERS — grouped under "Filters" (closed, but clickable)
    # =============================================
    filter_names = []

    if ec_near:
        lines.append("select EC_Near_Membrane, {} and chain {} and resi {}".format(
            pdb_name, ch, resi_str(ec_near)))
        filter_names.append("EC_Near_Membrane")
        lines.append("")

    if dist_only:
        lines.append("select Distant_Buried, {} and chain {} and resi {}".format(
            pdb_name, ch, resi_str(dist_only)))
        filter_names.append("Distant_Buried")
        lines.append("")

    if surf_only:
        lines.append("select Exposed_Not_Conserved, {} and chain {} and resi {}".format(
            pdb_name, ch, resi_str(surf_only)))
        filter_names.append("Exposed_Not_Conserved")
        lines.append("")

    if conserved_surf:
        lines.append("select Conserved_Not_Patched, {} and chain {} and resi {}".format(
            pdb_name, ch, resi_str(conserved_surf)))
        filter_names.append("Conserved_Not_Patched")
        lines.append("")

    if all_dist_pass:
        lines.append("select Cumul_Distance_Pass, {} and chain {} and resi {}".format(
            pdb_name, ch, resi_str(all_dist_pass)))
        filter_names.append("Cumul_Distance_Pass")
        lines.append("")

    if all_surf_pass:
        lines.append("select Cumul_Surface_Pass, {} and chain {} and resi {}".format(
            pdb_name, ch, resi_str(all_surf_pass)))
        filter_names.append("Cumul_Surface_Pass")
        lines.append("")

    if all_cons_pass:
        lines.append("select Cumul_Conservation_Pass, {} and chain {} and resi {}".format(
            pdb_name, ch, resi_str(all_cons_pass)))
        filter_names.append("Cumul_Conservation_Pass")
        lines.append("")

    if filter_names:
        lines.append("group Filters, {}".format(" ".join(filter_names)))
        lines.append("group Filters, close")
        lines.append("")

    # Membrane bilayer CGO disks
    if membrane:
        pdb_file = Path(run_dir) / pdb_subdir / ".{}_epitope.pdb".format(gene)
        if pdb_rotated:
            # PDB coordinates are rotated: ecto-axis = Y-up, anchor at origin.
            # Draw bilayer horizontally at Y = -half_thickness.
            ht = membrane.membrane_half_thickness
            lines.extend(generate_membrane_cgo_pml(
                membrane, str(pdb_file), ch,
                normal_override=np.array([0.0, 1.0, 0.0]),
                center_override=np.array([0.0, -ht, 0.0]),
            ))
        else:
            lines.extend(generate_membrane_cgo_pml(
                membrane, str(pdb_file), ch,
            ))

    # Final orientation
    if pdb_rotated:
        # Coordinates already rotated so ecto-axis = Y-up.
        # Identity view matrix: screen-right = X, screen-up = Y.
        lines.extend([
            "# --- Ectodomain-axis-aligned view (coordinates rotated) ---",
            "deselect",
            "set_view (\\",
            "    1.000000, 0.000000, 0.000000,\\",
            "    0.000000, 1.000000, 0.000000,\\",
            "    0.000000, 0.000000, 1.000000,\\",
            "    0.000000, 0.000000, -400.000000,\\",
            "    0.00, 0.00, 0.00,\\",
            "    100.000000, 900.000000, -20.000000)",
            "",
            "zoom",
            "set ray_opaque_background, on",
            "bg_color black",
            "",
        ])
    else:
        lines.extend([
            "# --- Final display settings ---",
            "deselect",
            "orient {} and chain {}".format(pdb_name, ch),
            "zoom {} and chain {}".format(pdb_name, ch),
            "set ray_opaque_background, on",
            "bg_color black",
            "",
        ])

    with open(pml_path, "w") as f:
        f.write("\n".join(lines))

    logger.info("  Wrote PyMOL script: %s", pml_path.name)


# ---------------------------------------------------------------------------
# Annotated sequence export
# ---------------------------------------------------------------------------

def export_annotated_sequences(run_dir, target, membrane, spatial_filter,
                               surface_analysis, conservation_result,
                               specificity_result, scores):
    """
    Write annotated FASTA and per-residue CSV for a target.
    """
    uid = target.uniprot_id
    gene = target.gene_name.lower()
    seq_dir = Path(run_dir) / "Annotated Sequences"

    # --- FASTA ---
    fasta_path = seq_dir / "{}_epitope.fasta".format(gene)
    patch_ranges = []
    for s in scores:
        patch_ranges.append(s.residue_range)

    header = ">{} | {} | epitope_patches: [{}]".format(
        target.gene_name, uid,
        "] [".join(patch_ranges) if patch_ranges else "none",
    )
    if conservation_result:
        header += " | cyno_identity: {:.1f}%".format(
            conservation_result.overall_identity * 100
        )

    with open(fasta_path, "w") as f:
        f.write(header + "\n")
        # Write sequence in 80-char lines
        seq = target.sequence
        for i in range(0, len(seq), 80):
            f.write(seq[i:i+80] + "\n")

    logger.info("  Wrote annotated FASTA: %s", fasta_path.name)

    # --- Per-residue CSV ---
    csv_path = seq_dir / "{}_residue_table.csv".format(gene)

    # Build patch membership lookup
    patch_membership = {}
    for s in scores:
        for resnum in s.patch.residue_numbers:
            patch_membership[resnum] = s.patch_id

    fieldnames = [
        "residue_num", "aa", "topology", "distance_from_membrane_A",
        "sasa_A2", "relative_sasa", "in_patch", "patch_id",
        "cyno_conserved", "human_specific", "epitope_score",
    ]

    with open(csv_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for i, aa in enumerate(target.sequence, 1):
            topo = membrane.residue_topology.get(i, "") if membrane else ""
            dist = spatial_filter.residue_distances.get(i, 0.0) if spatial_filter else 0.0
            sasa = surface_analysis.residue_sasa.get(i, 0.0) if surface_analysis else 0.0
            rel_sasa = surface_analysis.residue_relative_sasa.get(i, 0.0) if surface_analysis else 0.0
            patch_id = patch_membership.get(i, "")
            conserved = ""
            if conservation_result:
                c = conservation_result.residue_conservation.get(i)
                if c is not None:
                    conserved = "yes" if c else "no"

            specific = ""
            if specificity_result and specificity_result.residue_specificity:
                s_val = specificity_result.residue_specificity.get(i)
                if s_val is True:
                    specific = "yes"
                elif s_val is False:
                    specific = "no"

            # Find epitope score for this residue
            ep_score = ""
            for s in scores:
                if i in s.patch.residue_numbers:
                    ep_score = round(s.composite_score, 4)
                    break

            writer.writerow({
                "residue_num": i,
                "aa": aa,
                "topology": topo,
                "distance_from_membrane_A": round(dist, 1) if dist else "",
                "sasa_A2": round(sasa, 1) if sasa else "",
                "relative_sasa": round(rel_sasa, 3) if rel_sasa else "",
                "in_patch": "yes" if i in patch_membership else "no",
                "patch_id": patch_id,
                "cyno_conserved": conserved,
                "human_specific": specific,
                "epitope_score": ep_score,
            })

    logger.info("  Wrote per-residue CSV: %s", csv_path.name)


# ---------------------------------------------------------------------------
# JSON annotation export
# ---------------------------------------------------------------------------

def export_annotation_json(run_dir, target, structure, membrane, spatial_filter,
                           surface_analysis, conservation_result, specificity_result,
                           scores, metric):
    """
    Write full annotation JSON for programmatic reuse.
    """
    json_path = Path(run_dir) / "Annotations" / "{}_annotation.json".format(
        target.gene_name.lower()
    )

    data = {
        "target": {
            "gene_name": target.gene_name,
            "uniprot_id": target.uniprot_id,
            "protein_name": target.protein_name,
            "sequence_length": target.sequence_length,
            "cyno_uniprot_id": target.cyno_uniprot_id,
        },
        "structure": {
            "source": structure.source if structure else None,
            "pdb_id": structure.pdb_id if structure else None,
            "method": structure.method if structure else None,
            "resolution": structure.resolution if structure else None,
            "chain_id": structure.chain_id if structure else None,
        } if structure else None,
        "membrane": {
            "topology_type": membrane.topology_type if membrane else None,
            "tm_segments": membrane.tm_segments if membrane else [],
            "source": membrane.source if membrane else None,
        } if membrane else None,
        "spatial_filter": {
            "min_distance_threshold": spatial_filter.min_distance_threshold if spatial_filter else None,
            "total_extracellular": spatial_filter.total_extracellular if spatial_filter else 0,
            "total_qualifying": spatial_filter.total_qualifying if spatial_filter else 0,
            "max_distance": spatial_filter.max_distance if spatial_filter else 0,
        } if spatial_filter else None,
        "epitope_scores": [
            {
                "patch_id": s.patch_id,
                "rank": s.rank,
                "composite_score": round(s.composite_score, 4),
                "patch_area_a2": round(s.patch_area_a2, 1),
                "n_residues": s.n_residues,
                "residue_range": s.residue_range,
                "residues": s.patch.residue_numbers,
                "avg_distance_a": round(s.avg_distance_a, 1),
                "cyno_identity": round(s.cyno_identity, 4),
                "max_off_target_identity": round(s.max_off_target_identity, 4),
                "component_scores": {
                    "area": round(s.area_score, 4),
                    "distance": round(s.distance_score, 4),
                    "conservation": round(s.conservation_score, 4),
                    "specificity": round(s.specificity_score, 4),
                    "accessibility": round(s.accessibility_score, 4),
                },
            }
            for s in scores
        ],
        "metric": metric,
    }

    with open(json_path, "w") as f:
        json.dump(data, f, indent=2)
    logger.info("  Wrote annotation JSON: %s", json_path.name)


# ---------------------------------------------------------------------------
# Manifest
# ---------------------------------------------------------------------------

def export_manifest(run_dir, targets, parameters):
    """
    Write input manifest for reproducibility.
    """
    logs_dir = Path(run_dir) / "Supplementary Files" / "Logs"
    logs_dir.mkdir(parents=True, exist_ok=True)
    manifest_path = logs_dir / "input_manifest.json"
    data = {
        "pipeline": "epitope_pipeline",
        "version": "0.1.0",
        "date": date.today().isoformat(),
        "targets": [
            {
                "identifier": t.gene_name,
                "uniprot_id": t.uniprot_id,
                "sequence_length": t.sequence_length,
            }
            for t in targets
        ],
        "parameters": parameters,
    }

    with open(manifest_path, "w") as f:
        json.dump(data, f, indent=2)
    logger.info("  Wrote input manifest: %s", manifest_path.name)
